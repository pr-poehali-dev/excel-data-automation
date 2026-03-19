"""
Загрузка и парсинг Excel/CSV файлов.
Принимает файл в base64, возвращает данные в виде таблицы (заголовки + строки).
"""

import json
import base64
import io


def handler(event: dict, context) -> dict:
    cors = {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type",
    }

    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": cors, "body": ""}

    try:
        body = json.loads(event.get("body") or "{}")
        file_b64 = body.get("file")
        filename = body.get("filename", "file.xlsx")
        sheet_index = int(body.get("sheetIndex", 0))
        header_row = int(body.get("headerRow", 0))
        encoding = body.get("encoding", "utf-8")
        delimiter = body.get("delimiter", ",")

        if not file_b64:
            return {"statusCode": 400, "headers": cors, "body": json.dumps({"error": "Файл не передан"})}

        file_bytes = base64.b64decode(file_b64)
        ext = filename.rsplit(".", 1)[-1].lower()

        if ext == "csv":
            rows, headers, sheet_names = _parse_csv(file_bytes, encoding, delimiter, header_row)
        elif ext in ("xlsx", "xls"):
            rows, headers, sheet_names = _parse_excel(file_bytes, sheet_index, header_row)
        else:
            return {"statusCode": 400, "headers": cors, "body": json.dumps({"error": f"Неподдерживаемый формат: .{ext}"})}

        return {
            "statusCode": 200,
            "headers": cors,
            "body": json.dumps({
                "headers": headers,
                "rows": rows,
                "totalRows": len(rows),
                "sheetNames": sheet_names,
                "filename": filename,
            }, ensure_ascii=False),
        }

    except Exception as e:
        return {"statusCode": 500, "headers": cors, "body": json.dumps({"error": str(e)})}


def _parse_excel(file_bytes: bytes, sheet_index: int, header_row: int):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames
    ws = wb.worksheets[min(sheet_index, len(wb.worksheets) - 1)]

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([_cell_value(c) for c in row])

    if not all_rows:
        return [], [], sheet_names

    headers = [str(h) if h is not None else f"Колонка {i+1}" for i, h in enumerate(all_rows[header_row])]
    rows = []
    for raw in all_rows[header_row + 1:]:
        if any(c is not None for c in raw):
            row_dict = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
            rows.append(row_dict)

    return rows, headers, sheet_names


def _parse_csv(file_bytes: bytes, encoding: str, delimiter: str, header_row: int):
    import csv
    text = file_bytes.decode(encoding, errors="replace")
    reader = list(csv.reader(io.StringIO(text), delimiter=delimiter))

    if not reader:
        return [], [], ["CSV"]

    headers = [str(h).strip() or f"Колонка {i+1}" for i, h in enumerate(reader[header_row])]
    rows = []
    for raw in reader[header_row + 1:]:
        if any(c.strip() for c in raw):
            row_dict = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))}
            rows.append(row_dict)

    return rows, headers, ["CSV"]


def _cell_value(val):
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)
